#!/usr/bin/python3
# -*- coding=utf-8 -*-
import openpyxl


def xlsx(path):
    """
    调整xlsx 格式
    :param path: 要调用的 xlsx 文件路径
    """
    # from openpyxl.styles import colors, fills, Font, PatternFill

    workbook1 = openpyxl.load_workbook(path)
    worksheet1 = workbook1['port']

    font = openpyxl.styles.Font(name=u'微软雅黑', size=10, color='FF0000')
    alig = openpyxl.styles.Alignment(horizontal="left", vertical='center', wrap_text=True)  # wrap_text 自动和行

    side = openpyxl.styles.Side(style='thin', color='000000')
    border = openpyxl.styles.Border(left=side, right=side, top=side, bottom=side)

    patt = openpyxl.styles.PatternFill(fill_type=None, fgColor='FF0000', bgColor='1E90FF')  # start_color 背景颜色

    for cell in worksheet1['1']:
        # 给首行添加字体和对齐
        cell.font = font
        cell.alignment = alig
        # 给首行添加填充
        cell.fill = patt

    # 给所有单元格加边框
    for row in worksheet1.iter_rows():
        for cell in row:
            cell.border = border

    # 设置行高、列宽
    for row in range(1, worksheet1.max_row+1):
        # 首行20
        if row == 1:
            worksheet1.row_dimensions[row].height = 25
        # 其他行15
        else:
            worksheet1.row_dimensions[row].height = 15

    # 第一列宽度15
    worksheet1.column_dimensions['A'].width = 10
    worksheet1.column_dimensions['I'].width = 25
    worksheet1.column_dimensions['J'].width = 25
    # for col in range(1, worksheet1.max_column + 1):
    #     worksheet1.column_dimensions[openpyxl.utils.get_column.letter(1)].width = 15
    workbook1.save(path)
