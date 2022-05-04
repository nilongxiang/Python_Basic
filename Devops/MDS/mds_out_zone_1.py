#!/usr/bin/python3
# -*- coding=utf-8 -*-
import openpyxl
import re
import os

f_zone = open('zone.txt', 'w+')


def get_zone(name):
    keyStart = '`show zoneset active vsan 1-4093`'
    KeyMiddle = 'zoneset name zoneset3000 vsan 3000'
    # KeyMiddle = 'zoneset name Zoneset3001 vsan 3001'
    keyEnd = '`show zoneset vsan 1-4093`'
    # 2个子组，返回一个元组
    regex = re.compile(keyStart + '(.*?)' + KeyMiddle + '(.*?)' + keyEnd, re.S)
    result = regex.findall(name.read())[0][1]
    for i in result:
        f_zone.write(i)
    f_zone.write('\n')  # 写入3行空行，便于append_zone函数退出循环
    f_zone.write('\n')
    f_zone.write('\n')
    f_zone.close()


def append_zone(sto_name):
    with open('zone.txt') as f_zone1:
        # f_zone.readline()
        # f_zone.seek(0)
        num = 3   # 因为表头为2行，后续循环range从0开始计数
        list_zone = ['start']  # 定义初始list_zone不为空，因为f_zone1第一行为空
        list_host = []  # 定义每个zone里主机wwpn
        list_storage = []  # 定义每个zone里存储wwpn
        while True:
            cont = f_zone1.readline()
            if re.fullmatch('\\s+', cont):  # 如果为空行,则表示取完一次数据,可以执行操作, 第一行不能为空;
                if not list_zone:  # 如果列表也为空,则表示数据读完了,结束循环
                    break
                # 第一步：将每个zone信息输出成列表
                # 查看list1内容
                # print(list_zone)

                # 第二步：获取list_zone里wwpn
                wwpn = re.findall('[0-9a-z]{2}:[0-9a-z]{2}:[0-9a-z]{2}:[0-9a-z]{2}:'
                                  '[0-9a-z]{2}:[0-9a-z]{2}:[0-9a-z]{2}:[0-9a-z]{2}', str(list_zone))
                # print(wwpn)

                # 第三步：将wwpn分别写入list_storage，list_host列表
                for i in wwpn:
                    if i in storage_wwpn.values():
                        list_storage.append(i)
                    else:
                        list_host.append(i)

                # 第四步：判断list_storage里wwpn品牌，如果有需要的品牌写入excel
                for j in list_storage:
                    r = re.findall('[a-zA-Z]{2,6}[0-9]{0,4}',
                                   list(storage_wwpn.keys())[list(storage_wwpn.values()).index(j)])
                    # 这是关键，匹配到目标存储再执行后续往excel表写的动作
                    if r == [sto_name]:
                        # 取到list_storage个数，依次循环写入excel
                        for k in range(len(list_storage)):
                            # 1: 写入存储wwpn
                            worksheet1.cell(k + num, 8).value = list_storage[k]

                            # 2: 写入存储wwpn对应的品牌
                            # print(list_storage[k])
                            # print(list(storage_wwpn.values()).index(list_storage[k]))
                            vendor = re.findall('[a-zA-Z]{2,6}[0-9]{0,4}',
                                                list(storage_wwpn.keys())[
                                                    list(storage_wwpn.values()).index(list_storage[k])])
                            worksheet1.cell(k + num, 7).value = str(vendor)

                            # 3: 写入存储wwpn是否在线
                            for online1 in list_zone:
                                if re.findall(list_storage[k], online1):
                                    a = re.findall('\\*', online1)
                                    worksheet1.cell(k + num, 9).value = str(a)

                        # 4: 写入主机wwpn
                        for m in range(len(list_host)):
                            # print(m)
                            worksheet1.cell(m + num, 5).value = list_host[m]
                            # 5: 写入主机wwpn是否在线
                            for online2 in list_zone:
                                if re.findall(list_host[m], online2):
                                    b = re.findall('\\*', online2)
                                    worksheet1.cell(m + num, 6).value = str(b)

                            # 6: 写入host name
                            workbook2 = openpyxl.load_workbook(
                                r'C:\Users\Andy\PycharmProjects\Python基础\MDS\DC-CBS-MDS9509-1_端口信息表1.xlsx')
                            # workbook2 = openpyxl.load_workbook(
                            #     r'C:\Users\Andy\PycharmProjects\Python基础\MDS\DC-CBS-MDS9509-2_端口信息表1.xlsx')
                            worksheet2 = workbook2['port']
                            for cell in worksheet2['I']:
                                if cell.value == list_host[m]:
                                    # print(cell.row)
                                    # print(worksheet2.cell(cell.row, 10).value)
                                    worksheet1.cell(m + num, 4).value = worksheet2.cell(cell.row, 10).value
                                    break
                            else:
                                worksheet1.cell(m + num, 4).value = 'other_switch'

                            # 7: 写入Zone name
                            # 获取zone name
                            name = re.findall('zone name (.*?) vsan 3000', str(list_zone))
                            worksheet1.cell(m + num, 3).value = str(name)

                        # 将list_storage里元素个数家到num值里，便于后续excel写入
                        num += len(list_storage)
                        # print(num)

                        # 判断上述list_storage里是否有需要的存储品牌，匹配到任意一个跳出循环
                        break
                # 判断上述list_storage里是否有需要的存储品牌，如果没有，则打印没有该品牌存储
                # else:
                #     print("没有hds存储")

                # print(list_storage)
                # print(list_host)
                list_storage.clear()
                list_host.clear()
                list_zone.clear()
            else:
                list_zone.append(cont)
    # 8: 写入Zoneset name和Storage name
    worksheet1.cell(num - 1, 2).value = 'Zoneset3000'
    # worksheet1.cell(num - 1, 2).value = 'Zoneset3001'
    worksheet1.cell(num - 1, 1).value = sto_name


if __name__ == '__main__':
    # 切换9509-2交换机，需要修改6处，main函数3处，get_zone函数1处，append_zone函数2处
    f_name = open('95091.log', encoding='utf-8')
    f_path = 'DC-CBS-MDS9509-1_Zone.xlsx'
    # f_name = open('95092.log', encoding='utf-8')
    # f_path = 'DC-CBS-MDS9509-2_Zone.xlsx'
    storage_wwpn = {"hds_1": '50:06:0e:80:16:64:cd:00',
                    "hds_2": '50:06:0e:80:16:64:cd:02',
                    "hds_3": '50:06:0e:80:16:64:cd:20',
                    "hds_4": '50:06:0e:80:16:64:cd:22',
                    "hds_5": '50:06:0e:80:16:64:cd:40',
                    "hds_6": '50:06:0e:80:16:64:cd:42',
                    "hds_7": '50:06:0e:80:16:64:cd:60',
                    "hds_8": '50:06:0e:80:16:64:cd:62',
                    "hds_9": '50:06:0e:80:16:64:cd:01',
                    "hds_10": '50:06:0e:80:16:64:cd:03',
                    "hds_11": '50:06:0e:80:16:64:cd:21',
                    "hds_12": '50:06:0e:80:16:64:cd:23',
                    "hds_13": '50:06:0e:80:16:64:cd:41',
                    "hds_14": '50:06:0e:80:16:64:cd:43',
                    "hds_15": '50:06:0e:80:16:64:cd:61',
                    "hds_16": '50:06:0e:80:16:64:cd:63',
                    "hp1_1": '20:12:00:02:ac:00:6c:82',
                    "hp1_2": '21:12:00:02:ac:00:6c:82',
                    "hp2_1": '21:12:00:02:ac:01:2f:13',
                    "hp2_2": '20:11:00:02:ac:01:2f:13',
                    "Dorado5000_1": '20:82:80:b5:75:c3:74:59',
                    "Dorado5000_2": '20:92:80:b5:75:c3:74:59',
                    "Dorado5000_3": '20:90:80:b5:75:c3:74:59',
                    "Dorado5000_4": '20:80:80:b5:75:c3:74:59',
                    "Unity_1": '50:06:01:64:47:e0:26:44',
                    "Unity_2": '50:06:01:6d:47:e0:26:44',
                    "emc_1": '50:00:09:74:08:4e:55:18',
                    "emc_2": '50:00:09:74:08:4e:55:19',
                    "emc_3": '50:00:09:74:08:4e:55:58',
                    "emc_4": '50:00:09:74:08:4e:55:59',
                    "emc_5": '50:00:09:74:08:4e:55:98',
                    "emc_6": '50:00:09:74:08:4e:55:99',
                    "emc_7": '50:00:09:74:08:4e:55:d8',
                    "emc_8": '50:00:09:74:08:4e:55:1c',
                    "emc_9": '50:00:09:74:08:4e:55:1d',
                    "emc_10": '50:00:09:74:08:4e:55:5c',
                    "emc_11": '50:00:09:74:08:4e:55:5d',
                    "emc_12": '50:00:09:74:08:4e:55:9c',
                    "emc_13": '50:00:09:74:08:4e:55:9d',
                    "emc_14": '50:00:09:74:08:4e:55:dc'}

    # storage_wwpn = {"hds_1": '50:06:0e:80:16:64:cd:11',
    #                 "hds_2": '50:06:0e:80:16:64:cd:13',
    #                 "hds_3": '50:06:0e:80:16:64:cd:31',
    #                 "hds_4": '50:06:0e:80:16:64:cd:33',
    #                 "hds_5": '50:06:0e:80:16:64:cd:51',
    #                 "hds_6": '50:06:0e:80:16:64:cd:53',
    #                 "hds_7": '50:06:0e:80:16:64:cd:71',
    #                 "hds_8": '50:06:0e:80:16:64:cd:73',
    #                 "hds_9": '50:06:0e:80:16:64:cd:10',
    #                 "hds_10": '50:06:0e:80:16:64:cd:12',
    #                 "hds_11": '50:06:0e:80:16:64:cd:30',
    #                 "hds_12": '50:06:0e:80:16:64:cd:32',
    #                 "hds_13": '50:06:0e:80:16:64:cd:50',
    #                 "hds_14": '50:06:0e:80:16:64:cd:52',
    #                 "hds_15": '50:06:0e:80:16:64:cd:70',
    #                 "hds_16": '50:06:0e:80:16:64:cd:72',
    #                 "hp1_1": '20:11:00:02:ac:00:6c:82',
    #                 "hp1_2": '21:11:00:02:ac:00:6c:82',
    #                 "hp2_1": '21:11:00:02:ac:01:2f:13',
    #                 "hp2_2": '20:12:00:02:ac:01:2f:13',
    #                 "Dorado5000_1": '20:93:80:b5:75:c3:74:59',
    #                 "Dorado5000_2": '20:83:80:b5:75:c3:74:59',
    #                 "Dorado5000_3": '20:91:80:b5:75:c3:74:59',
    #                 "Dorado5000_4": '20:81:80:b5:75:c3:74:59',
    #                 "Unity_1": '50:06:01:65:47:e0:26:44',
    #                 "Unity_2": '50:06:01:6c:47:e0:26:44',
    #                 "emc_1": '50:00:09:74:08:4e:55:24',
    #                 "emc_2": '50:00:09:74:08:4e:55:25',
    #                 "emc_3": '50:00:09:74:08:4e:55:64',
    #                 "emc_4": '50:00:09:74:08:4e:55:65',
    #                 "emc_5": '50:00:09:74:08:4e:55:a4',
    #                 "emc_6": '50:00:09:74:08:4e:55:a5',
    #                 "emc_7": '50:00:09:74:08:4e:55:e4',
    #                 "emc_8": '50:00:09:74:08:4e:55:20',
    #                 "emc_9": '50:00:09:74:08:4e:55:21',
    #                 "emc_10": '50:00:09:74:08:4e:55:60',
    #                 "emc_11": '50:00:09:74:08:4e:55:61',
    #                 "emc_12": '50:00:09:74:08:4e:55:a0',
    #                 "emc_13": '50:00:09:74:08:4e:55:a1',
    #                 "emc_14": '50:00:09:74:08:4e:55:e0'}

    # 第一步：从show tech-support detail里提取zoneset active
    get_zone(f_name)

    # 第二步：创建每个存储的sheet
    workbook1 = openpyxl.Workbook()
    for storage_name in ['hp1', 'hp2', 'Unity', 'Dorado5000', 'emc', 'hds']:
        worksheet1 = workbook1.create_sheet(storage_name)
        worksheet1['A1'] = 'Storage'
        worksheet1['B1'] = 'Zoneset name'
        worksheet1['C1'] = 'Zone name'
        worksheet1['D1'] = 'Zone host member'
        worksheet1['G1'] = 'Zone storage member'
        worksheet1['D2'] = 'host name'
        worksheet1['E2'] = 'host wwpn'
        worksheet1['F2'] = 'connection\n（* indicated as online）'
        worksheet1['G2'] = 'storage name'
        worksheet1['H2'] = 'storage wwpn'
        worksheet1['I2'] = 'connection\n（* indicated as online）'

        # 第三步：插入zone数据
        append_zone(storage_name)

    # 第四步：执行完数据插入，保存Zone.xlsx
    workbook1.save('DC-CBS-MDS9509-1_Zone.xlsx')
    # workbook1.save('DC-CBS-MDS9509-2_Zone.xlsx')

    # 第五步：删除临时文件
    os.chdir(r'C:\Users\Andy\PycharmProjects\Python基础\MDS')
    os.remove(r'C:\Users\Andy\PycharmProjects\Python基础\MDS\zone.txt')

