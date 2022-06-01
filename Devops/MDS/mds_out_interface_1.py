#!/usr/bin/python3
# -*- coding=utf-8 -*-
import openpyxl
import re
import xlsx_format_interface
import os

f_port = open('interface.txt', 'w+')
f_wwpn = open('flogi.txt', 'w+')
f_description = open('description.txt', 'w+')


def get_interface(txt):
    """
    获取show interface brief输出信息
    :param txt: 原始log name
    """
    keyStart = '`show interface brief`'
    keyEnd = '`show interface`'
    pat = re.compile(keyStart + '(.*?)' + keyEnd, re.S)
    result = pat.findall(txt.read())
    for i in result:
        f_port.write(i)


def get_wwpn(txt):
    """
    获取show flogi database输出信息
    :param txt: 原始log name
    """
    keyStart = '`show flogi database`'
    keyEnd = 'Total number of flogi'
    pat = re.compile(keyStart + '(.*?)' + keyEnd, re.S)
    # 文件指针回到开始
    txt.seek(0)
    result = pat.findall(txt.read())
    for i in result:
        f_wwpn.write(i)


def get_description(txt):
    """
    获取show running-config输出信息
    :param txt: 原始log name
    """
    keyStart = '`show running-config.txt`'
    keyEnd = '`show startup-config.txt`'
    pat = re.compile(keyStart + '(.*?)' + keyEnd, re.S)
    # 文件指针回到开始
    txt.seek(0)
    result = pat.findall(txt.read())
    for i in result:
        f_description.write(i)


def out_port_xlsx():
    """
    根据get_interface函数信息，输出端口表
    """
    f_port.seek(0)
    for i in f_port.readlines():
        # # 过滤包含trunking或up的行
        # if re.match("fc.*trunking.*", i) or re.match("fc.*up.*", i):
        if re.search('^fc.*(trunking|up).*', i):
            # 依据空格切片并转成list，用于excel写入
            i_list = list(i.split())
            # 弹出最后一列
            i_list.pop()
            worksheet1.append(i_list)
    f_port.close()


def append_wwpn_xlsx():
    """
    在输出端口表里追加 wwpn信息
    """
    f_wwpn.seek(0)
    for i in f_wwpn.readlines():
        if re.match("fc.*", i):
            i_list = list(i.split())
            # print(i_list)
            for cell in worksheet1["A"]:
                if i_list[0] == cell.value:
                    # print(i_list[0], i_list[3])
                    worksheet1.cell(row=cell.row, column=9, value=i_list[3])
    f_wwpn.close()


def append_description_xlsx():
    """
    在输出端口表里追加 description信息
    """
    f_description.seek(0)
    while True:
        line = f_description.readline()
        # 匹配到interface GigabitEthernet，跳出循环
        if re.match('interface GigabitEthernet', line):
            break
        # 匹配到interface fc行时，继续使用readline读取下一行并赋值给a
        elif re.match('interface fc', line):
            a = f_description.readline()
            # 匹配a中description
            if re.findall('description', a):
                # print(line.split()[1], ' '.join(a.split()[2:]))
                for cell in worksheet1["A"]:
                    # 将line使用空格切片取到端口名与xlsx里比较
                    if line.split()[1] == cell.value:
                        # value值为将a使用空格切片,取到索引2及之后数据,使用空格将2及之后数据拼接
                        worksheet1.cell(row=cell.row, column=10, value=' '.join(a.split()[2:]))
    for cell in worksheet1["J"]:
        if cell.value is None:
            cell.value = 'Not description'
    f_description.close()


def out_xlsx(name, path):
    """
    输出 f_name 端口信息表.xlsx
    :param name: 原始 log name
    :param path: 输出 xlsx文件路径
    """
    # 第一步：创建xlsx实例
    global workbook1
    workbook1 = openpyxl.Workbook()
    # 创建port sheet
    global worksheet1
    worksheet1 = workbook1.create_sheet('port', 0)

    # 定义sheet表头
    sheet_title = ["槽位/端口", "VSAN", "Admin Mode", "Admin Trunk mode", "运行状态", "SFP", "运行状态", "速率", "WWPN", "端口描述"]
    worksheet1.append(sheet_title)

    # 第二步：分别获取interface、wwpn、description数据并关闭f_name
    get_interface(name)
    get_wwpn(name)
    get_description(name)
    name.close()

    # 第三步：输出show interface信息
    out_port_xlsx()

    # 第四步：增加WWPN
    append_wwpn_xlsx()

    # 第五步：增加interface description
    append_description_xlsx()

    # 第六步：保存xlsx并添加格式
    workbook1.save(path)
    xlsx_format_interface.xlsx(path)


if __name__ == '__main__':
    # 切换9509-2交换机，需要修改main函数1处即可
    f_path = 'DC-CBS-MDS9509-1_端口信息表1.xlsx'
    f_name = open('95091.log', encoding='utf-8')
    # f_path = 'DC-CBS-MDS9509-2_端口信息表1.xlsx'
    # f_name = open('95092.log', encoding='utf-8')
    out_xlsx(f_name, f_path)
    os.chdir(r'E:\05-04 笔记\PycharmProject_Windows\MDS')
    os.remove(r'E:\05-04 笔记\PycharmProject_Windows\MDS\interface.txt')
    os.remove(r'E:\05-04 笔记\PycharmProject_Windows\MDS\flogi.txt')
    os.remove(r'description.txt')
